# -*- coding: utf-8 -*-
"""
TC 입력 JSON -> TC 시트 xlsx (5시트 자기완결 산출물)

파이프라인에서의 위치
--------------------
  test-case/{프로젝트}-tc-input-v{X.Y}.json   (설계 원본 = 정본)
    │  이 스크립트
    ▼
  test-case/{프로젝트}-tc-v{X.Y}.xlsx          (파생 = 배포물)

생성 시트
--------
  Test Case      한 행 = 한 TN = 한 스텝. 절차의 "1. / 2."를 스텝으로 분해하고
                 문체를 Test-Step '~한다' / Expected-Result '~된다'로 정규화
  Summary        케이스 단위 자동 집계(COUNTIFS). 기준 골격 버전을 C4에 자동 기입
  이슈 관리 시트   결함 기록(JIRA 미사용, 내장 운영). 드롭다운은 '목록' 시트 참조
  목록           드롭다운 참조 목록의 정본. 숨기지 않고 맨 뒤 배치
  명세서          구조 규칙의 정본 시트

규칙 정본
--------
  구조·컬럼   project-process/rules/tc-sheet-format.md + 이 산출물의 '명세서' 시트
  서식        design-template/xlsx-design-guide.md
              (헤더 #1F2A44/#F3F3F3 · 데이터 #FFFFFF/#000000 · 표 테두리 #999999
               · 실행 채움 셀 #FFF9C4 · 머리 영역 #BFBFBF/#DEFFE4 · 뎁스 채움 §5
               · 맑은 고딕 · 날짜 YYYY-MM-DD)

openpyxl 함정 3종 (2026-08-02 디버깅으로 확인 — Excel '복구' 프롬프트의 원인)
--------------------------------------------------------------------------
  1) freeze_panes=None 만으로 틀 고정을 풀면 pane을 참조하는 selection이 남아
     파일이 깨진다. sheet_view.selection 도 기본값으로 초기화해야 한다
  2) DataValidation.formula1 에 '=' 접두를 붙이면 안 된다 (openpyxl이 그대로 기록)
     교차 시트 참조는 정의된 이름(defined name)을 쓰는 편이 호환에 안전하다
  3) '='로 시작하는 설명 문구는 수식으로 저장되어 파일을 깨뜨린다.
     명세서의 수식 원형은 '=' 없이 적고 본문에서 안내한다

사용법:
    python build_tc_template_xlsx.py input.json -o out.xlsx [--issues issues.json]
    (생성 후 Excel로 열어 경고 없이 열리는지 확인할 것 — 함정 재발 감시)

이슈 기록은 별도 파일(--issues)이 정본입니다. TC 설계 원본은 확정되면 고정되지만 이슈는
실행하면서 계속 늘고 상태가 바뀌므로, 한 파일에 두면 이슈 갱신 때마다 설계 원본이 변경되어
확정 시점이 흐려집니다. 스키마는 {"issues": [{project, no, status, summary, description,
priority, frequency, affectedVersion, fixedVersion, resolution, environment, label,
reporter, assignee, watcher, attachment, sprint, createdDay, updatedDay,
statusChangedDay}]} 이며 키 순서는 ISSUE_KEYS가 정본입니다.

입력 JSON 스키마
----------------
{
  "title": "Test Case Template",
  "project": "miyonchat",                    # 이슈 ID 접두 · 목록 시트 기본값
  "tree_version": "{프로젝트}-tree-vX.Y",     # Summary C4 자동 기입 (마스터는 플레이스홀더)
  "platforms": ["Web", "And", "iOS"],
  "d1_order": ["앱 진입", ...],               # 1-Depth 표시 순서 (생략 시 등장 순)
  "env": {"OS": ["", "", ""], ...},          # 상단 환경 블록 기본값
  "lists": {"레이블": [...], ...},            # 목록 시트 덮어쓰기 (선택)
  "vt_note": {"루브릭": "…"},                 # 검증유형 문구 덮어쓰기 (선택)
                                             #   기본값은 rules/verification-types.md,
                                             #   프로젝트가 design/에서 다르게 정했을 때만 씁니다
  "issue_samples": [[...20개 값...]],         # 이슈 시트 예시 행 (선택)
  "tcs": [
    ["TC-XXX-001",["1-Depth","2-Depth","3-Depth"],"케이스",
     "사전조건","1. 절차\\n2. 절차","기대결과 문장. 여러 문장 가능.",
     "결정적|확률적|루브릭|금칙","자동화 전용|공통|사람 전용","High|Medium|Low",
     "선행 TC ID 또는 -","대상 서비스","Note(수행 안내)"]
    # 8번째 값(실행 주체)은 2026-08-03 신설이며 생략하면 「공통」으로 읽습니다(11필드 호환)
  ]
}

뎁스 열은 쓰는 만큼만 만듭니다 (2026-08-03 개정)
------------------------------------------------
  시트는 최대 7뎁스를 지원하지만, 트리 깊이는 프로젝트마다 다르고 같은 프로젝트 안에서도
  가지마다 다릅니다. 그래서 **경로를 배열로 받아** 실제로 쓰인 최대 깊이를 세고, 그만큼만
  열을 만듭니다. 케이스명은 경로 다음 칸에 놓입니다.

  다 만든 뒤 남는 열을 지우는 방식은 쓰지 않습니다 — Summary가 Test Case의 열 문자를
  직접 가리키므로(COUNTIFS), 열을 지우면 Priority·Result 열이 앞으로 밀려 집계가 엉뚱한
  열을 셉니다. openpyxl은 다른 시트의 수식을 따라 고쳐 주지 않습니다.

  **옛 형식(d1·d2·d3 세 칸 고정, 13필드)도 그대로 읽습니다** — 두 번째 값이 배열이면
  새 형식, 문자열이면 옛 형식으로 봅니다.
"""
import argparse
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from norm import to_step, to_expected, split_steps, split_expected  # noqa: E402,F401

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.views import Selection
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

# ── 서식 토큰 (xlsx-design-guide.md) ─────────────────────────
FONT = "맑은 고딕"
NAVY, SUBH = "1F2A44", "44546A"
HEADER_TEXT = "F3F3F3"
LABEL_FILL = "EDEFF4"
INPUT_FILL = "FFF9C4"        # 실행 채움 셀 — 설계 시점에는 비어 있는 칸
NOTE_TEXT = "555555"
BORDER_RGB = "999999"
HEADBAND_FILL = "BFBFBF"     # 머리 영역의 회색 면 — 헤더 + 환경 블록의 좌·우 구간
ENVBAND_FILL = "F9D7BE"      # 머리 영역에서 Total Result·Result 열이 지나는 구간

HDR = PatternFill("solid", fgColor=NAVY)
SUB = PatternFill("solid", fgColor=SUBH)
LABEL = PatternFill("solid", fgColor=LABEL_FILL)
INPUT = PatternFill("solid", fgColor=INPUT_FILL)
WHITE = PatternFill("solid", fgColor="FFFFFF")
HEADBAND = PatternFill("solid", fgColor=HEADBAND_FILL)
ENVBAND = PatternFill("solid", fgColor=ENVBAND_FILL)

_side = Side(style="thin", color=BORDER_RGB)
BOX = Border(left=_side, right=_side, top=_side, bottom=_side)

# 상태 배색 (xlsx-design-guide.md §3) — 값 기반 조건부 서식.
# 빈 칸은 노란 채움(채울 자리)으로 남고, 값이 들어가면 상태색이 덮습니다.
GREEN = ("D1FAE5", "047857")
RED = ("FEE2E2", "B91C1C")
PURPLE = ("EDE9FE", "6D28D9")
WHITE_C = ("FFFFFF", "000000")
ORANGE = ("FEF3C7", "B45309")
BLUE = ("DBEAFE", "1D4ED8")
GRAY_C = ("E5E7EB", "4B5563")

STATUS_COLORS = {
    "Pass": GREEN, "Resolved": GREEN,
    "Fail": RED, "High": RED,
    "Blocked": PURPLE,
    "Open": WHITE_C, "Reopen": WHITE_C,
    "Medium": ORANGE,
    "In Progress": BLUE,
    "Closed": GRAY_C, "Low": GRAY_C, "NI": GRAY_C,
}


def paint_status(ws, ref, values, size=9):
    """ref 범위의 셀을 값에 따라 상태색으로 칠한다."""
    for v in values:
        fill, fontc = STATUS_COLORS[v]
        ws.conditional_formatting.add(
            ref,
            CellIsRule(operator="equal", formula=[f'"{v}"'],
                       fill=PatternFill("solid", fgColor=fill),
                       font=Font(name=FONT, size=size, bold=True, color=fontc)))


# 검증유형별 판정 규칙 문구 — 워크스페이스 기본값(rules/verification-types.md).
# 프로젝트가 design/에서 다른 값을 확정했다면 입력 json의 vt_note로 덮어씁니다.
VT_NOTE = {
    "결정적": "결정적 · 1회 실행, 기대값 불일치 시 FAIL",
    "확률적": "확률적 · 20회 반복(지표성 30~50회), 명시 임계 미달 시 FAIL",
    "루브릭": "루브릭 · 5점 채점, 평가자 2인 또는 심판모델 3회, 합격선 4점",
    "금칙": "금칙 · 우회 변형 포함 20~40회 시도, 1건이라도 발생 시 FAIL",
}

# 실행 주체 3종 (2026-08-03 확정) — 케이스마다 「누가 수행하는가」.
# 판정은 순서대로: ①사람이 화면만 봐서 판정 불가 → 자동화 전용
#                 ②자동화가 판정 기준을 못 세움 → 사람 전용  ③그 외 → 공통(기본값)
# 반복 횟수는 판정 근거가 아닙니다 — "사람이 1회 봐도 의미가 있는가"로 갈립니다.
EXEC_TYPES = ["자동화 전용", "공통", "사람 전용"]
EXEC_DEFAULT = "공통"

DEFAULT_LISTS = {
    "프로젝트": ["{프로젝트}"],
    "실행 주체": EXEC_TYPES,
    "이슈 상태": ["Open", "In Progress", "Resolved", "Reopen", "Closed"],
    "우선순위": ["High", "Medium", "Low"],
    "빈도": ["Always", "Often", "Sometimes", "Once"],
    "버전(영향/수정 공용)": ["{테스트환경}_{개발목표버전}_{스프린트}_{확인버전}"],
    "해결책": ["Fixed", "Won't Do", "Won't Fix", "Duplicate", "Incomplete",
              "Cannot Reproduce", "Done", "Declined"],
    "환경": ["PC웹", "Android", "iOS"],
    "레이블": ["{기능 트리 1-Depth 영역명}"],
    "보고자": ["{작성자}"],
    "담당자": ["{담당자}"],
    "스프린트": ["{테스트환경}_{개발목표버전}_{스프린트}"],
}
LIST_NAME_MAP = {
    "프로젝트": "list_project", "실행 주체": "list_exec",
    "이슈 상태": "list_status", "우선순위": "list_priority",
    "빈도": "list_frequency", "버전(영향/수정 공용)": "list_version",
    "해결책": "list_resolution", "환경": "list_env", "레이블": "list_label",
    "보고자": "list_reporter", "담당자": "list_assignee", "스프린트": "list_sprint",
}
ISSUE_HEADERS = ["프로젝트 ID", "Issue No.", "이슈 상태", "요약(Summary)", "설명",
                 "우선순위", "빈도", "영향 받는 버전", "수정 버전", "해결책", "환경",
                 "레이블", "보고자", "담당자", "관측자", "첨부파일", "스프린트",
                 "이슈 등록일", "이슈 최종 수정일자", "이슈 상태 최종 변경일자"]
# 이슈 json의 키 → 컬럼 순서. 사람이 고치는 파일이라 배열이 아니라 객체로 씁니다
ISSUE_KEYS = ["project", "no", "status", "summary", "description",
              "priority", "frequency", "affectedVersion", "fixedVersion", "resolution",
              "environment", "label", "reporter", "assignee", "watcher", "attachment",
              "sprint", "createdDay", "updatedDay", "statusChangedDay"]
ISSUE_WIDTHS = [12, 13, 12, 46, 50, 9, 10, 17, 17, 16, 8, 16, 9, 10, 12, 10, 13, 12, 14, 15]
# 드롭다운을 붙일 컬럼 -> 목록 이름
ISSUE_DV = {"B": "프로젝트", "D": "이슈 상태", "G": "우선순위", "H": "빈도",
            "I": "버전(영향/수정 공용)", "J": "버전(영향/수정 공용)", "K": "해결책",
            "L": "환경", "M": "레이블", "N": "보고자", "O": "담당자", "R": "스프린트"}


def hfont(size=9, color=HEADER_TEXT):
    return Font(name=FONT, size=size, bold=True, color=color)


def dfont(size=9, bold=False, color="000000"):
    return Font(name=FONT, size=size, bold=bold, color=color)


def unfreeze(ws):
    """틀 고정 해제 — selection까지 초기화하지 않으면 파일이 깨진다(함정 1)."""
    ws.freeze_panes = None
    ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]


MAX_DEPTH = 7          # 시트가 지원하는 뎁스 열의 최대 수


# ── 뎁스 색상 배정 (xlsx-design-guide.md §5) ─────────────────
# 규칙(2026-08-03 사용자 확정): ①빈칸은 칠하지 않음 ②같은 텍스트는 같은 색
# ③다른 텍스트는 다른 색 ④연한 계열 ⑤인접한 다른 텍스트와 비슷한 계열 회피 ⑥색 최소화
#
# ③이 색 수를 텍스트 수로 고정하므로 ⑥은 "색을 적게 쓰라"가 아니라 "색상환을 고르게
# 나눠 쓰라"로 구현됩니다 — 텍스트 수만큼 등간격 색을 만들면 그 이상 촘촘해지지 않습니다.
# ⑤는 실제 인접 관계를 세어 그리디로 배정합니다: 시트에서 위아래(같은 열의 값 전환)와
# 좌우(같은 행의 이웃 뎁스)로 붙는 쌍을 모아, 각 텍스트에 **이미 배정된 이웃과의 색상환
# 거리가 최대가 되는** 색을 줍니다. 이웃이 많아 여유가 없으면 남은 것 중 최선을 씁니다(⑤ 단서).
#
# 노랑 대역(45~70°)은 건너뜁니다 — 이 시트에서 노랑은 "실행 단계에 채워지는 칸"이라는
# 뜻을 이미 갖고 있어, 뎁스에 쓰면 색의 의미 체계가 충돌합니다.
YELLOW_BAND = (45, 70)
PASTEL_S = 0.42            # 연한 계열 — 글씨(검정)를 가리지 않는 범위
PASTEL_L = 0.87


def _hsl_to_hex(h, s, light):
    c = (1 - abs(2 * light - 1)) * s
    x = c * (1 - abs((h / 60.0) % 2 - 1))
    m = light - c / 2
    r, g, b = [(c, x, 0), (x, c, 0), (0, c, x),
               (0, x, c), (x, 0, c), (c, 0, x)][int(h // 60) % 6]
    return "".join("%02X" % round((v + m) * 255) for v in (r, g, b))


def _hue_slots(n):
    """노랑 대역을 뺀 색상환을 n등분한 색조 목록."""
    span = 360 - (YELLOW_BAND[1] - YELLOW_BAND[0])
    out = []
    for i in range(n):
        h = (i + 0.5) * span / n
        out.append(h if h < YELLOW_BAND[0] else h + (YELLOW_BAND[1] - YELLOW_BAND[0]))
    return out


def _circ(a, b):
    d = abs(a - b) % 360
    return min(d, 360 - d)


def depth_palette(paths):
    """뎁스 경로 목록(시트에 찍히는 순서)을 받아 {텍스트: 채움색}을 만든다."""
    seen, adj = [], {}
    prev = []
    for p in paths:
        for v in p:
            if v and v not in adj:
                adj[v] = set()
                seen.append(v)
        for a, b in zip(p, p[1:]):            # 좌우 이웃 — 같은 행의 다음 뎁스
            if a and b and a != b:
                adj[a].add(b)
                adj[b].add(a)
        for a, b in zip(prev, p):             # 위아래 이웃 — 같은 열의 값 전환
            if a and b and a != b:
                adj[a].add(b)
                adj[b].add(a)
        prev = p

    slots = _hue_slots(len(seen))
    free = set(range(len(slots)))
    got = {}
    # 이웃이 많은 것부터 — 제약이 큰 쪽을 먼저 놓아야 뒤에서 몰리지 않는다
    for t in sorted(seen, key=lambda x: (-len(adj[x]), seen.index(x))):
        near = [slots[got[o]] for o in adj[t] if o in got]
        pick = (max(free, key=lambda i: (min(_circ(slots[i], h) for h in near), -i))
                if near else min(free))
        got[t] = pick
        free.discard(pick)

    # 교환 개선 — 그리디는 마지막에 놓이는 것들이 남은 색을 받게 되어 한두 쌍이 붙는다.
    # 가장 가까운 쌍을 골라 다른 텍스트와 색을 맞바꿔 보고, 전체 최솟값이 커지면 채택한다.
    def worst_of(t, table):
        near = [slots[table[o]] for o in adj[t]]
        return min((_circ(slots[table[t]], h) for h in near), default=360)

    for _ in range(200):
        low = min(seen, key=lambda t: worst_of(t, got))
        base = worst_of(low, got)
        best = None
        for other in seen:
            if other is low:
                continue
            trial = dict(got)
            trial[low], trial[other] = got[other], got[low]
            gain = min(worst_of(low, trial), worst_of(other, trial))
            if gain > base and (best is None or gain > best[0]):
                best = (gain, other)
        if best is None:
            break
        got[low], got[best[1]] = got[best[1]], got[low]

    return {t: PatternFill("solid", fgColor=_hsl_to_hex(slots[i], PASTEL_S, PASTEL_L))
            for t, i in got.items()}


def normalize_tc(row):
    """TC 한 줄을 dict로 정규화한다 — 새 형식(경로 배열)과 옛 형식(d1·d2·d3)을 함께 받는다.

    두 번째 값이 배열이면 새 형식이다. 옛 형식은 빈 뎁스를 떨어내 같은 모양으로 만든다.
    """
    if isinstance(row[1], list):
        # 실행 주체(exec)는 2026-08-03 신설 — 없으면 기본값 「공통」으로 읽습니다
        if len(row) >= 12:
            tid, path, case, pre, steps, exp, vt, ex, prio, par, target, note = row[:12]
        else:
            tid, path, case, pre, steps, exp, vt, prio, par, target, note = row
            ex = EXEC_DEFAULT
    else:
        tid, d1, d2, d3, case, pre, steps, exp, vt, prio, par, target, note = row
        path = [d for d in (d1, d2, d3) if d]
        ex = EXEC_DEFAULT
    return {"id": tid, "path": list(path), "case": case, "pre": pre, "steps": steps,
            "exp": exp, "vt": vt, "exec": ex or EXEC_DEFAULT, "prio": prio,
            "par": par, "target": target, "note": note}


def depth_layout(tcs):
    """실제로 쓰인 뎁스 열의 수 — 경로 최대 길이 + 케이스명 한 칸.

    다 만들고 지우는 대신 처음부터 필요한 만큼만 만든다. Summary가 Test Case의 열 문자를
    가리키므로, 나중에 지우면 Priority·Result 열이 밀려 집계가 어긋난다.
    """
    longest = max((len(t["path"]) for t in tcs), default=1)
    return min(longest + 1, MAX_DEPTH)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("-o", "--output", default="tc.xlsx")
    ap.add_argument("--issues", help="이슈 기록 json (없으면 입력의 issue_samples를 씁니다)")
    args = ap.parse_args()

    CFG = json.load(open(args.input, encoding="utf-8"))
    TCS = [normalize_tc(t) for t in CFG["tcs"]]
    PLATFORMS = CFG.get("platforms", ["Web", "And", "iOS"])
    # 2026-08-03 개정으로 시트 제목은 title이 아니라 「기준 골격 버전 + {프로젝트} TC」
    # 두 토막이 됩니다. title은 파일 설명용으로만 남아 시트에 찍히지 않습니다.
    PROJECT = CFG.get("project", "{프로젝트}")
    TREE_VERSION = CFG.get("tree_version", "{프로젝트}-tree-vX.Y")
    ENV = CFG.get("env", {k: [""] * len(PLATFORMS) for k in
                          ["OS", "단말", "버전", "작업자 이름", "작업 시작일"]})
    ORDER = CFG.get("d1_order") or list(dict.fromkeys(t["path"][0] for t in TCS))

    LISTS = dict(DEFAULT_LISTS)
    LISTS["프로젝트"] = [PROJECT]
    for k, v in (CFG.get("lists") or {}).items():
        LISTS[k] = v

    # 검증유형 문구 — 프로젝트가 design/에서 확정한 값이 있으면 덮어쓴다
    vt_note = dict(VT_NOTE)
    vt_note.update(CFG.get("vt_note") or {})

    # 선행 관계 -> 실행 단계
    PAR = {t["id"]: t["par"] for t in TCS}
    layer = {}

    def depth(tid):
        if tid in layer:
            return layer[tid]
        p = PAR.get(tid, "-")
        layer[tid] = 1 if p in ("-", "", None) else depth(p) + 1
        return layer[tid]

    # ── 컬럼 레이아웃 ────────────────────────────────────────
    # 뎁스 수를 먼저 정하고 그 뒤 열 위치를 계산한다 — 지도를 그리고 채우는 순서다.
    # 사람이 왼쪽부터 읽고 실행하는 구간 뒤에, 기계·참조용 값(TC ID·검증유형)을 오른쪽
    # 끝 참조 블록으로 뺍니다(2026-08-03 확정). 조인 키와 분류는 실행 동선에 끼어들 이유가
    # 없고, 이슈를 등록하며 TC ID를 옮겨 적을 때 Issue No.·Comment 바로 옆이 편합니다.
    ND = depth_layout(TCS)
    NP = len(PLATFORMS)
    DEPTH_W = [15, 14, 15, 20, 11, 11, 11]

    n = 3                              # C열부터 시작 (A=여백, B=No)
    def take(k=1):
        nonlocal n
        got = [get_column_letter(n + i) for i in range(k)]
        n += k
        return got if k > 1 else got[0]

    DEPTH_COLS = take(ND)
    if ND == 1:
        DEPTH_COLS = [DEPTH_COLS]
    PRE, TN_COL, STEP, EXPECT, EXEC_COL, PRIO_COL = (take(), take(), take(),
                                                     take(), take(), take())
    TOTAL_COLS = take(NP) if NP > 1 else [take()]
    RESULT_COLS = take(NP) if NP > 1 else [take()]
    ISSUE_COL, COMMENT, NOTE, TCID_COL, VT_COL = (take(), take(), take(), take(), take())
    LAST = VT_COL   # Test Case Edit 열은 2026-08-03 사용자 결정으로 제거 — 이력은 json+git이 담당

    ENV_ROW = 5                        # 환경 블록 첫 행 (2행 제목 · 3~4행 헤더 다음)
    DATA_ROW = ENV_ROW + len(ENV)      # 데이터 첫 행 = 틀 고정 경계

    base = [("A", "", 3), ("B", "No", 6)]
    base += [(DEPTH_COLS[i], "%d-Depth" % (i + 1), DEPTH_W[i]) for i in range(ND)]
    base += [(PRE, "Pre-Condition", 26), (TN_COL, "TN", 5),
             (STEP, "Test-Step", 34), (EXPECT, "Expected-Result", 46),
             (EXEC_COL, "실행 주체", 11), (PRIO_COL, "Priority", 9)]
    tail = [(ISSUE_COL, "Issue No.", 11), (COMMENT, "Comment", 34),
            (NOTE, "Note", 30), (TCID_COL, "TC ID", 13), (VT_COL, "검증유형", 9)]
    ALL_COLS = [b[0] for b in base[1:]] + TOTAL_COLS + RESULT_COLS + [t[0] for t in tail]

    wb = Workbook()

    # ══ Test Case ═══════════════════════════════════════════
    ws = wb.active
    ws.title = "Test Case"
    ws.sheet_view.showGridLines = False

    # ── 머리 영역 (2026-08-03 확정) ────────────────────────────
    # A열과 1행은 값도 색도 두지 않는다 — 다른 어떤 규칙보다 우선하는 여백 규칙이다.
    #   2행      제목 두 토막 (남색·흰 글씨·좌측 정렬) — 왼쪽=기준 골격 버전, 오른쪽={프로젝트} TC
    #   3~4행    컬럼 헤더. Total Result·Result만 2단(3행 묶음명 / 4행 플랫폼)이고
    #            나머지 열은 3:4를 세로 병합한다
    #   5행~     환경 블록 (라벨은 Priority 열, 값은 Result 열)
    # 회색 면(#BFBFBF)은 3행부터 데이터 직전까지 No~Priority·Issue No.~Note를 덮고,
    # Total Result·Result 열은 같은 구간을 연녹색(#DEFFE4)으로 덮어 실행 입력 자리를 가른다.
    HEAD_L = f"B2:{DEPTH_COLS[-1]}2"
    HEAD_R = f"{PRE}2:{LAST}2"
    for rng, text in ((HEAD_L, TREE_VERSION), (HEAD_R, f"{PROJECT} TC")):
        ws.merge_cells(rng)
        x = ws[rng.split(":")[0]]
        x.value = text
        x.font = hfont(11)
        x.fill = HDR
        x.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    for col, label, _w in base[1:] + tail:
        ws.merge_cells(f"{col}3:{col}4")
        x = ws[f"{col}3"]
        x.value = label
        x.font = hfont()
        x.fill = HEADBAND
        x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells(f"{TOTAL_COLS[0]}3:{TOTAL_COLS[-1]}3")
    ws[f"{TOTAL_COLS[0]}3"].value = "Total Result"
    ws.merge_cells(f"{RESULT_COLS[0]}3:{RESULT_COLS[-1]}3")
    ws[f"{RESULT_COLS[0]}3"].value = "Result"
    for cc in (TOTAL_COLS[0], RESULT_COLS[0]):
        ws[f"{cc}3"].font = hfont(color="000000")
        ws[f"{cc}3"].fill = ENVBAND
        ws[f"{cc}3"].alignment = Alignment(horizontal="center", vertical="center")
    for col, lab in list(zip(TOTAL_COLS, PLATFORMS)) + list(zip(RESULT_COLS, PLATFORMS)):
        x = ws[f"{col}4"]
        x.value = lab
        x.font = hfont(color="000000")
        x.fill = ENVBAND
        x.alignment = Alignment(horizontal="center", vertical="center")
    for r in (3, 4):
        for col in ALL_COLS:
            ws[f"{col}{r}"].border = BOX
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 16

    # 환경 블록 — 라벨(Priority 열)과 값(Result 열). 라벨 왼쪽 열들은 행마다 가로 병합해
    # 빈 칸이 격자로 쪼개져 보이지 않게 한다
    for i, (label, vals) in enumerate(ENV.items()):
        r = ENV_ROW + i
        ws.merge_cells(f"B{r}:{PRIO_COL}{r}")
        lc = ws[f"B{r}"]
        lc.value = label
        lc.font = hfont()
        lc.alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells(f"{TOTAL_COLS[0]}{r}:{TOTAL_COLS[-1]}{r}")
        for col, v in zip(RESULT_COLS, list(vals) + [""] * NP):
            x = ws[f"{col}{r}"]
            x.value = v
            x.font = dfont(bold=True)
            x.alignment = Alignment(horizontal="center", vertical="center")

    # 회색·연녹 면 — 3행부터 데이터 직전까지. 값이 있는 헤더 칸도 같은 면을 쓴다
    for r in range(3, DATA_ROW):
        for col in ALL_COLS:
            x = ws[f"{col}{r}"]
            x.fill = ENVBAND if col in TOTAL_COLS + RESULT_COLS else HEADBAND
    for r in range(ENV_ROW, DATA_ROW):        # 환경 값 = 실행 채움 셀 — 띠 위에 노랑
        for col in RESULT_COLS:
            ws[f"{col}{r}"].fill = INPUT
            ws[f"{col}{r}"].border = BOX
    for col in ALL_COLS:                      # 헤더 두 줄만 테두리를 되살린다
        for r in (3, 4):
            ws[f"{col}{r}"].border = BOX

    # 안내 문구 — 환경 블록 오른쪽(Issue No.~Note)의 빈 면을 한 덩어리로 묶어 넣는다.
    # 시트를 처음 여는 사람이 색 규약을 여기서 읽습니다(정본은 명세서 시트).
    ws.merge_cells(f"{ISSUE_COL}{ENV_ROW}:{LAST}{DATA_ROW - 1}")
    nc = ws[f"{ISSUE_COL}{ENV_ROW}"]
    nc.value = ("※ 노란색 셀은 실행 단계에서 채워집니다 — 환경 정보 / Result / Issue No. / "
                "Comment\n※ 수행 전 참고사항은 전부 Note에 있습니다 — 설계 데이터는 노란 칸에 "
                "두지 않습니다")
    nc.font = Font(name=FONT, size=8, color="333333")
    nc.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")

    # 데이터 행
    idx = {d: [] for d in ORDER}
    for t in TCS:
        idx.setdefault(t["path"][0], []).append(t)

    # 뎁스 색상 — 시트에 찍히는 순서(1-Depth 표시 순서 → 케이스 순)대로 모아 배정한다.
    # 경로와 케이스명을 함께 넣는다 — 뎁스 열에 놓이는 값은 전부 같은 규칙을 받는다.
    palette = depth_palette([(t["path"] + [t["case"]])[:ND]
                             for d1 in ORDER for t in idx.get(d1, [])])

    row = DATA_ROW
    for d1 in ORDER:
        for tc in idx.get(d1, []):
            tid, case, prio = tc["id"], tc["case"], tc["prio"]
            target, note, vt = tc["target"], tc["note"], tc["vt"]
            # 경로 + 케이스명을 뎁스 열에 깔고, 남는 칸은 비운다
            cells = (tc["path"] + [case])[:ND]
            cells += [""] * (ND - len(cells))
            stepl = split_steps(tc["steps"])
            expl = split_expected(tc["exp"])
            aligned = len(expl) == len(stepl)
            first = row
            blocks = []   # 스텝별 (시작행, 끝행) — 병합 범위 계산용
            for i, st in enumerate(stepl, start=1):
                is_last = i == len(stepl)
                bstart = row
                ws[f"B{row}"] = f"=ROW()-{DATA_ROW - 1}"
                for ci, v in enumerate(cells):
                    ws[f"{DEPTH_COLS[ci]}{row}"] = v
                ws[f"{PRE}{row}"] = tc["pre"] if i == 1 else ""
                ws[f"{TN_COL}{row}"] = i
                ws[f"{STEP}{row}"] = to_step(st)
                ws[f"{EXPECT}{row}"] = (expl[i - 1] if aligned
                                        else (expl[0] if (is_last and expl) else ""))
                # 케이스 단위 값은 행마다 반복 — 행 단위 집계와 필터가 성립하려면 값이
                # 모든 행에 있어야 합니다(병합하면 필터에 첫 행만 걸립니다)
                ws[f"{PRIO_COL}{row}"] = prio
                ws[f"{EXEC_COL}{row}"] = tc["exec"]
                ws[f"{TCID_COL}{row}"] = tid
                ws[f"{VT_COL}{row}"] = vt
                if i == 1 and note:
                    # Note는 사람이 읽을 안내만 담는다 — 함께 볼 것 / 수행 방법 / 조건 만드는 법.
                    # 케이스 메타(TC ID·검증유형)는 2026-08-03 개정으로 각자 열을 갖는다
                    ws[f"{NOTE}{row}"] = note
                row += 1
                if is_last and not aligned:
                    # 기대결과 추가 행 — No·Depth·Priority는 행마다 기재(행 단위 집계·필터의 전제),
                    # TN·Test-Step은 비워 두고 아래에서 스텝 범위로 병합
                    for extra in expl[1:]:
                        ws[f"B{row}"] = f"=ROW()-{DATA_ROW - 1}"
                        for ci, v in enumerate(cells):
                            ws[f"{DEPTH_COLS[ci]}{row}"] = v
                        ws[f"{EXPECT}{row}"] = extra
                        ws[f"{PRIO_COL}{row}"] = prio
                        ws[f"{EXEC_COL}{row}"] = tc["exec"]
                        ws[f"{TCID_COL}{row}"] = tid
                        ws[f"{VT_COL}{row}"] = vt
                        row += 1
                blocks.append((bstart, row - 1))
            for r in range(first, row):
                for col in ALL_COLS:
                    x = ws[f"{col}{r}"]
                    x.border = BOX
                    x.font = dfont()
                    center = (col in ("B", TN_COL, PRIO_COL, EXEC_COL, TCID_COL, VT_COL)
                              or col in TOTAL_COLS or col in RESULT_COLS)
                    x.alignment = Alignment(horizontal="center" if center else "left",
                                            vertical="center", wrap_text=True)
                    x.fill = INPUT if col in RESULT_COLS + [ISSUE_COL, COMMENT] else WHITE
                # 뎁스 채움 — 값이 있는 칸만(경로·케이스명 모두). 빈칸은 흰 배경
                for ci in range(ND):
                    x = ws[f"{DEPTH_COLS[ci]}{r}"]
                    if x.value:
                        x.fill = palette.get(x.value, WHITE)
                ws[f"C{r}"].font = dfont(bold=True)
                ws[f"{TN_COL}{r}"].font = dfont(bold=True)
                ws.row_dimensions[r].height = 30
            # 셀 병합 (2026-08-03 확정) — 스텝 범위: TN·Test-Step은 한 동작 = 한 덩어리로,
            # TN 1 스텝의 Pre-Condition은 같은 조건이 걸리는 범위로 묶는다 (케이스 전체 병합 금지)
            for bi, (b0, b1) in enumerate(blocks):
                if b1 > b0:
                    ws.merge_cells(f"{TN_COL}{b0}:{TN_COL}{b1}")
                    ws.merge_cells(f"{STEP}{b0}:{STEP}{b1}")
                    ws.merge_cells(f"{NOTE}{b0}:{NOTE}{b1}")
                    if bi == 0:
                        ws.merge_cells(f"{PRE}{b0}:{PRE}{b1}")
            # Total Result — 병합하지 않고 데이터 한 행씩 세운다(2026-08-03 확정).
            # 같은 플랫폼에 단말이 여럿일 때 그 행의 단말 결과를 하나로 요약하는 자리이며,
            # 스텝 여러 개를 한 칸으로 묶으면 어느 행의 판정인지 읽히지 않습니다.
            for tcol, rcol in zip(TOTAL_COLS, RESULT_COLS):
                for r in range(first, row):
                    cell = ws[f"{tcol}{r}"]
                    rng = f"{rcol}{r}:{rcol}{r}"
                    cell.value = (f'=IF(COUNTIF({rng},"Fail")>0,"Fail",'
                                  f'IF(COUNTIF({rng},"Blocked")>0,"Blocked",'
                                  f'IF(COUNTIF({rng},"NI")>0,"NI",'
                                  f'IF(COUNTIF({rng},"Pass")>0,"Pass",""))))')
                    cell.font = dfont(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    last_row = row - 1
    dv = DataValidation(type="list", formula1='"Pass,Fail,NI,Blocked"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{RESULT_COLS[0]}{DATA_ROW}:{RESULT_COLS[-1]}500")
    dvp = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(dvp)
    dvp.add(f"{PRIO_COL}{DATA_ROW}:{PRIO_COL}500")
    for col, vals in ((EXEC_COL, EXEC_TYPES), (VT_COL, list(VT_NOTE))):
        d = DataValidation(type="list", formula1='"%s"' % ",".join(vals), allow_blank=True)
        ws.add_data_validation(d)
        d.add(f"{col}{DATA_ROW}:{col}500")

    # 상태 배색 — 실행 결과(Result)와 케이스 판정(Total Result), 우선순위
    paint_status(ws, f"{RESULT_COLS[0]}{DATA_ROW}:{RESULT_COLS[-1]}500", ["Pass", "Fail", "NI", "Blocked"])
    paint_status(ws, f"{TOTAL_COLS[0]}{DATA_ROW}:{TOTAL_COLS[-1]}500", ["Pass", "Fail", "NI", "Blocked"])
    paint_status(ws, f"{PRIO_COL}{DATA_ROW}:{PRIO_COL}500", ["High", "Medium", "Low"])

    for col, _l, w in base + tail:
        ws.column_dimensions[col].width = w
    for col in TOTAL_COLS + RESULT_COLS:
        ws.column_dimensions[col].width = 7

    ws.auto_filter.ref = f"B4:{LAST}{max(last_row, DATA_ROW)}"

    # 틀 고정 — 데이터 첫 행 위. 스크롤해도 제목·헤더·환경 블록이 남는다.
    # A열·1행은 값도 색도 없는 여백이므로 고정 대상에서 뺀다(B{DATA_ROW} 기준 고정).
    ws.freeze_panes = f"A{DATA_ROW}"

    # ══ Summary ═════════════════════════════════════════════
    s2 = wb.create_sheet("Summary")
    s2.sheet_view.showGridLines = False
    s2["B2"] = "커버리지 요약"
    s2["B2"].font = dfont(12, bold=True)
    s2.row_dimensions[2].height = 19
    s2["B3"] = "Test Case 시트를 참조하는 수식으로 자동 집계됩니다. 행을 추가하면 아래 범위를 넓히세요."
    s2["B3"].font = Font(name=FONT, size=8, color=NOTE_TEXT)
    s2["B4"] = "기준 골격 버전"
    s2["B4"].font = dfont(bold=True)
    s2["B4"].fill = LABEL
    s2["B4"].border = BOX
    s2.merge_cells("C4:E4")
    # 자동 기입 — 사람이 옮겨 적지 않으므로 TC 세트와 골격 버전이 어긋날 수 없다
    s2["C4"] = TREE_VERSION
    s2["C4"].font = dfont()
    s2["C4"].fill = WHITE
    for cc in ("C4", "D4", "E4"):
        s2[cc].border = BOX

    s2.merge_cells("B5:B6")
    s2["B5"] = "1-Depth (영역)"
    s2.merge_cells("C5:C6")
    s2["C5"] = "TC 수"
    s2.merge_cells("D5:F5")
    s2["D5"] = "Priority"
    for i, p in enumerate(PLATFORMS):
        c0 = 7 + i * 4
        s2.merge_cells(start_row=5, start_column=c0, end_row=5, end_column=c0 + 3)
        s2.cell(5, c0, p)
    for cc in ["B5", "C5", "D5"] + [get_column_letter(7 + i * 4) + "5" for i in range(NP)]:
        s2[cc].font = hfont()
        s2[cc].fill = HDR
        s2[cc].alignment = Alignment(horizontal="center", vertical="center")
    sub6 = ["High", "Medium", "Low"] + ["Pass", "Fail", "Blocked", "Pass율"] * NP
    for i, lab in enumerate(sub6):
        x = s2.cell(6, 4 + i, lab)
        x.font = hfont()
        x.fill = SUB
        x.alignment = Alignment(horizontal="center", vertical="center")
    ncols = 3 + 3 + 4 * NP           # B..(마지막)
    for r in (5, 6):
        for ci in range(2, 2 + ncols):
            s2.cell(r, ci).border = BOX

    for i, d1 in enumerate(ORDER):
        r = 7 + i
        a = s2.cell(r, 2, d1)
        a.font = dfont(bold=True)
        a.border = BOX
        a.fill = WHITE
        tcr = "'Test Case'!"
        # 행 수 기준 집계 — TC 수와 결과 집계의 단위를 스텝 행으로 맞춘다
        f = {3: f'=COUNTIF({tcr}$C${DATA_ROW}:$C$500,$B{r})'}
        for j, pr in enumerate(["High", "Medium", "Low"]):
            f[4 + j] = (f'=COUNTIFS({tcr}$C${DATA_ROW}:$C$500,$B{r},'
                        f'{tcr}${PRIO_COL}${DATA_ROW}:${PRIO_COL}$500,"{pr}")')
        for pi, rcol in enumerate(RESULT_COLS):
            c0 = 7 + pi * 4
            for j, st in enumerate(["Pass", "Fail", "Blocked"]):
                f[c0 + j] = (f'=COUNTIFS({tcr}$C${DATA_ROW}:$C$500,$B{r},'
                             f'{tcr}${rcol}${DATA_ROW}:${rcol}$500,"{st}")')
            pcol, fcol = get_column_letter(c0), get_column_letter(c0 + 1)
            f[c0 + 3] = f'=IF({pcol}{r}+{fcol}{r}=0,"",{pcol}{r}/({pcol}{r}+{fcol}{r}))'
        for ci, formula in f.items():
            x = s2.cell(r, ci, formula)
            x.font = dfont()
            x.fill = WHITE
            x.alignment = Alignment(horizontal="center")
            x.border = BOX
            if (ci - 7) % 4 == 3 and ci >= 7:
                x.number_format = "0.0%"

    tot = 7 + len(ORDER)
    tc0 = s2.cell(tot, 2, "합계")
    tc0.font = hfont()
    tc0.fill = SUB
    tc0.border = BOX
    for ci in range(3, 2 + ncols):
        col = get_column_letter(ci)
        if (ci - 7) % 4 == 3 and ci >= 7:
            pcol, fcol = get_column_letter(ci - 3), get_column_letter(ci - 2)
            v = f'=IF({pcol}{tot}+{fcol}{tot}=0,"",{pcol}{tot}/({pcol}{tot}+{fcol}{tot}))'
        else:
            v = f"=SUM({col}7:{col}{tot - 1})"
        x = s2.cell(tot, ci, v)
        x.font = hfont()
        x.fill = SUB
        x.border = BOX
        x.alignment = Alignment(horizontal="center")
        if (ci - 7) % 4 == 3 and ci >= 7:
            x.number_format = "0.0%"

    s2.cell(tot + 2, 2,
            "TC 수 = 스텝 행 수(1-Depth 기준) · Pass/Fail/Blocked도 같은 행 단위(Result 열) · "
            "Pass율 = Pass ÷ (Pass + Fail) — NI·Blocked는 분모에서 제외합니다"
            ).font = Font(name=FONT, size=8, color=NOTE_TEXT)
    s2.column_dimensions["A"].width = 3
    s2.column_dimensions["B"].width = 22
    for ci in range(3, 2 + ncols):
        s2.column_dimensions[get_column_letter(ci)].width = 9
    unfreeze(s2)

    # ══ 이슈 관리 시트 ═══════════════════════════════════════
    s3 = wb.create_sheet("이슈 관리 시트")
    s3.sheet_view.showGridLines = False
    for i, h in enumerate(ISSUE_HEADERS):
        x = s3.cell(2, 2 + i, h)
        x.font = hfont(11)
        x.fill = HDR
        x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        x.border = BOX
    # 이슈는 별도 파일이 정본입니다 — TC 설계 원본과 생명주기가 달라 파일을 나눕니다.
    # 파일이 없으면 입력의 issue_samples(마스터 템플릿용 예시)로 폴백합니다.
    if args.issues:
        rows = json.load(open(args.issues, encoding="utf-8")).get("issues", [])
        samples = [[str(r.get(k, "")) for k in ISSUE_KEYS] for r in rows]
    else:
        samples = CFG.get("issue_samples")
    if not samples:
        samples = [[
            PROJECT, f"{PROJECT}-1", "Open",
            "{화면} 진입 > {동작} 시, {결과} 안 됨",
            "Pre-condition: {사전 조건}\nReproduce Step: {재현 스텝}\n"
            "Actual Result: {실제 결과}\nExpected Result: {기대 결과}\nQA Comment: {참고}",
            "High", "Always", "", "", "", "PC웹", "", "", "", "", "", "",
            "YYYY-MM-DD", "YYYY-MM-DD", "YYYY-MM-DD"]]
    for r, rowv in enumerate(samples, start=3):
        for i, v in enumerate(rowv):
            x = s3.cell(r, 2 + i, v)
            x.font = dfont(11)
            x.fill = WHITE
            x.alignment = Alignment(vertical="top", wrap_text=(i in (3, 4)))
            x.border = BOX
    for i, w in enumerate(ISSUE_WIDTHS):
        s3.column_dimensions[get_column_letter(2 + i)].width = w
    s3.column_dimensions["A"].width = 3
    s3.row_dimensions[2].height = 30
    s3.freeze_panes = "A3"                       # 시트별 옵션: 헤더 행 고정
    s3.auto_filter.ref = f"B2:{get_column_letter(1 + len(ISSUE_HEADERS))}{2 + len(samples)}"

    # ══ 목록 ════════════════════════════════════════════════
    s4 = wb.create_sheet("목록")
    s4.sheet_view.showGridLines = False
    for li, (name, values) in enumerate(LISTS.items()):
        col = 2 + li
        x = s4.cell(2, col, name)
        x.font = hfont(11)
        x.fill = HDR
        x.alignment = Alignment(horizontal="center", vertical="center")
        x.border = BOX
        for r, v in enumerate(values, start=3):
            y = s4.cell(r, col, v)
            y.font = dfont(11)
            y.fill = WHITE
            y.border = BOX
        s4.column_dimensions[get_column_letter(col)].width = \
            max(14, max(len(str(v)) for v in values) + 4)
        # 정의된 이름 등록 — 드롭다운은 이 이름을 참조한다(함정 2)
        ref = f"목록!${get_column_letter(col)}$3:${get_column_letter(col)}${2 + len(values)}"
        dn = LIST_NAME_MAP[name]
        wb.defined_names[dn] = DefinedName(dn, attr_text=ref)
    s4.column_dimensions["A"].width = 3
    unfreeze(s4)

    for col, list_name in ISSUE_DV.items():
        d = DataValidation(type="list", formula1=LIST_NAME_MAP[list_name],
                           allow_blank=True, showDropDown=False)
        s3.add_data_validation(d)
        d.add(f"{col}3:{col}200")

    # 상태 배색 — 이슈 상태(D)와 우선순위(G)
    paint_status(s3, "D3:D200", ["Open", "In Progress", "Resolved", "Reopen", "Closed"], size=11)
    paint_status(s3, "G3:G200", ["High", "Medium", "Low"], size=11)
    # 목록 시트의 해당 열도 같은 색으로 — 값 체계를 색까지 함께 보여줍니다
    for list_name, ref in (("이슈 상태", "C3:C20"), ("우선순위", "D3:D20")):
        paint_status(s4, ref, LISTS[list_name], size=11)

    # ══ 명세서 ═══════════════════════════════════════════════
    build_spec_sheet(wb)

    wb.save(args.output)
    print(f"saved {args.output} | TC {len(TCS)} | rows {last_row - 10} | "
          f"platforms {PLATFORMS} | tree {TREE_VERSION}")


def build_spec_sheet(wb):
    ws = wb.create_sheet("명세서")
    ws.sheet_view.showGridLines = False

    def H(*vals):
        return ("head",) + vals

    def R(*vals):
        return ("row",) + vals

    SPEC = [
        ("title", "명세서 — 시트 규칙 정본"),
        ("note", "이 시트가 구조 규칙의 정본입니다. rules/tc-sheet-format.md와 짝으로 교차 "
                 "검증하며, 불일치 발견 시 임의 판단 없이 사용자에게 기준을 확인합니다. "
                 "서식(색·테두리·폰트)의 정본은 design-template/xlsx-design-guide.md입니다."),
        ("gap",),
        ("section", "1. 시트 구성"),
        H("시트", "역할"),
        R("Test Case", "TC 작성과 실행 기록. 노란 셀은 실행 단계에서 채워집니다"),
        R("Summary", "행 단위 자동 집계(1-Depth 기준). 기준 골격 버전(C4)은 생성 시 자동 기입됩니다"),
        R("이슈 관리 시트", "결함 기록(내장 운영, JIRA 미사용) — JIRA 이슈 등록·관리 방식을 "
                       "시트로 표현. Issue No.로 Test Case와 연결합니다. 정본은 별도 파일 "
                       "test-case/{프로젝트}-issues.json이며 이 시트는 그 파생입니다"),
        R("목록", "드롭다운 참조 목록의 정본. 숨기지 않고 맨 뒤에 둡니다"),
        R("명세서", "이 시트 — 구조 규칙의 정본"),
        ("gap",),
        ("section", "2. 컬럼 정의 (Test Case)"),
        H("컬럼", "내용", "채우는 규칙"),
        R("No", "행 번호", "자동 수식(ROW() 기준). 직접 입력하지 않습니다"),
        R("1~7-Depth", "기능 계층", "행마다 반복(기대결과 추가 행 포함 — 집계·필터의 전제). "
          "의미 있는 깊이까지만 쓰고 나머지는 빈칸. 병합하지 않습니다"),
        R("Pre-Condition", "사전조건",
          "TN 1행에 케이스 진입 조건. 중간 스텝에 앞 스텝이 만들지 못하는 새 상태(시간 경과·"
          "외부 변화 등)가 필요하면 그 행에도 적습니다 — 앞 스텝의 결과로 성립한 상태는 제외. "
          "같은 조건이 걸리는 스텝 행 범위는 세로 병합하되 케이스 전체 병합은 금지"),
        R("TN", "스텝 번호", "케이스마다 1부터. 새 케이스가 시작되면 1로 복귀. 한 스텝의 "
          "기대결과가 여러 행이면(TN 값 무관) Test-Step과 함께 그 범위를 세로 병합합니다 — "
          "반복 기재하면 독립 동작 여러 개로 오독됩니다"),
        R("Test-Step", "수행 동작", "「~한다」 체. TN과 같은 범위로 세로 병합"),
        R("Expected-Result", "기대 결과",
          "「~된다」 체. 판정 가능한 문장 — 임계·합격선·시도 횟수는 문장 안에 숫자로. "
          "한 행 = 한 판정이므로 병합하지 않습니다"),
        R("실행 주체", "누가 수행하는가",
          "행마다 반복(드롭다운 3종). 자동화 전용 = 사람이 화면만 봐서 판정 불가 / "
          "사람 전용 = 자동화가 판정 기준을 못 세움 / 공통 = 기본값. 반복 횟수는 판정 "
          "근거가 아니며 「사람이 1회 봐도 의미가 있는가」로 갈립니다"),
        R("Priority", "우선순위",
          "행마다 반복. High/Medium/Low — 케이스의 속성이지만 Summary가 행 단위로 "
          "집계하므로 행마다 값을 둡니다"),
        R("Total Result", "행 판정", "수식 열. 병합하지 않고 데이터 한 행씩 세웁니다 — "
          "같은 플랫폼에 단말이 여럿일 때 그 행의 단말 결과를 요약합니다. 직접 입력하지 않습니다"),
        R("Result", "스텝 실행 결과", "드롭다운 4종(아래 상태값 정의). 실행 단계에서 채움 — 노란 셀"),
        R("Issue No.", "관련 이슈",
          "해당 TC를 수행하는 과정에서 이슈가 발생하면, 먼저 이슈 관리 시트에 이슈를 등록하고 "
          "거기서 부여된 Issue No.를 이 칸에 적습니다. 실행 단계에서 채움 — 노란 셀"),
        R("Comment", "수행 중 기록",
          "TC를 수행하다 생긴 문제·관찰을 실행 단계에서 적습니다 — 노란 셀. 설계 시점에는 비어 있습니다"),
        R("Note", "수행 전 참고사항",
          "사람이 읽을 안내만 담습니다 — ①함께 볼 것 ②수행 방법 ③조건 만드는 법. "
          "케이스 메타는 2026-08-03 개정으로 TC ID·검증유형 열이 따로 갖습니다. 스텝 범위 세로 병합"),
        R("TC ID", "조인 키",
          "행마다 반복. 자동화 케이스명·추적 매트릭스·이슈 연결이 참조합니다. 사람의 실행 "
          "동선에서 빼려고 Note 우측 참조 블록에 둡니다"),
        R("검증유형", "판정 방식",
          "행마다 반복(드롭다운 4종). 반복 횟수는 시트에 적지 않습니다 — 수동 실행자가 "
          "자기에게 내려진 지시로 오독하기 때문입니다. 반복은 자동화가 수행하며, 사람은 "
          "어떤 유형이든 1회입니다. 확률적·루브릭을 사람이 1회 수행했다면 결과는 Pass가 "
          "아니라 관찰 기록이고, 판정은 자동화 반복이나 채점이 담당합니다"),
        ("gap",),
        ("section", "3. 상태값 정의"),
        H("상태", "정의", "Pass율 분모"),
        R("Pass", "성공", "포함"),
        R("Fail", "실패", "포함"),
        R("NI", "미구현이거나 스펙에 없어 실행 대상이 아님 (Not Implemented)", "제외"),
        R("Blocked", "기능은 구현됐으나 확인할 수 없는 상태 — 선행 케이스 Fail·환경 결함·데이터 준비 불가 등. 사유는 실행 과정에서 Comment에 적습니다", "제외"),
        ("note", "Pass율 = Pass ÷ (Pass + Fail). NI·Blocked를 분모에 넣지 않는 이유는 결함 "
                 "1건이 후속 Blocked 수만큼 중복 계상되는 것을 막기 위해서입니다."),
        ("gap",),
        ("section", "4. Total Result 규칙"),
        H("항목", "규칙"),
        R("우선순위", "Fail > Blocked > NI > Pass — 스텝 중 하나라도 Fail이면 케이스 Fail"),
        R("수식 원형",
          'IF(COUNTIF(범위,"Fail")>0,"Fail",IF(COUNTIF(범위,"Blocked")>0,"Blocked",'
          'IF(COUNTIF(범위,"NI")>0,"NI",IF(COUNTIF(범위,"Pass")>0,"Pass",""))))'
          ' — 셀에는 앞에 = 를 붙여 사용합니다. 아무것도 실행되지 않은 케이스는 빈칸입니다'),
        R("병합", "하지 않습니다 — 데이터 한 행씩 세웁니다. 같은 플랫폼에 단말이 여럿일 때 그 행의 단말 결과를 하나로 요약하는 자리입니다"),
        ("gap",),
        ("section", "5. TC ID 체계"),
        H("항목", "규칙"),
        R("형식", "TC-{영역코드}-{번호 3자리} — 예: TC-ENT-001"),
        R("번호", "영역 안에서만 증가합니다. 다른 영역에 케이스가 추가돼도 기존 ID가 흔들리지 않습니다"),
        R("영역코드", "기능 트리 1-Depth당 하나. 매핑표는 프로젝트 시트에서 정의합니다 (예: 앱 진입=ENT)"),
        R("기재 위치", "Note의 첫 토큰(케이스 메타의 맨 앞). 자동화 케이스명·추적 매트릭스·"
          "이슈 연결이 이 ID를 참조합니다. Comment는 설계 시점에 비어 있는 실행 기록 칸이므로 "
          "설계 데이터를 두지 않습니다"),
        ("gap",),
        ("section", "6. 플랫폼 열 규칙"),
        H("항목", "규칙"),
        R("기본값", "마스터는 Web / And / iOS 3열"),
        R("프로젝트 조정", "TC 설계 시작 시 대상 플랫폼을 사용자에게 재확인하고, Test Case의 "
                      "Total Result·Result 하위 열과 Summary의 플랫폼 열을 함께 조정합니다"),
        ("gap",),
        ("section", "7. 입력 규칙"),
        H("항목", "규칙"),
        R("실행 채움 셀(노랑)", "설계 시점에는 비어 있고 실행 단계에서 채워지는 칸 — 환경 블록·"
                        "Result·Issue No.·Comment. 채우는 주체가 사람이든 자동화든 설계자는 "
                        "건드리지 않습니다. 설계 데이터는 노란 칸에 두지 않습니다"),
        R("설계 셀", "Depth·절차·기대결과는 실행 중 수정하지 않습니다"),
        R("틀 고정", "데이터 첫 행 위. 스크롤해도 2행 제목·헤더·환경 블록이 남으며, "
                  "그 구간은 회색(#BFBFBF)과 Total/Result 띠(#F9D7BE)로 데이터와 갈라 둡니다. "
                  "A열과 1행은 값도 색도 두지 않습니다"),
        R("자동 필터", "헤더 행에 걸어 둡니다. 실행 주체·검증유형·Priority처럼 케이스 단위 값으로 "
                    "거르면 한 케이스의 행이 통째로 남거나 빠져 스텝 범위 병합이 깨지지 않습니다. "
                    "Result 같은 행 단위 값으로 거르면 케이스 중간이 잘려 병합이 어색해집니다"),
        R("뎁스 채움 색", "뎁스 열의 값 있는 셀만 텍스트별 고유색으로 채웁니다 — 같은 텍스트는 "
                     "같은 색, 인접한 다른 텍스트와는 색상환에서 멀어지게 배정. 케이스명 칸과 "
                     "빈칸은 흰 배경. 규칙 정본은 xlsx-design-guide.md §5"),
        R("기준 골격 버전", "Summary C4에 생성 스크립트가 자동 기입합니다 — 형식 {프로젝트}-tree-v{X.Y}. "
                      "TC 설계 입력(tc-input json)의 값을 그대로 쓰므로 사람이 옮겨 적지 않습니다"),
        ("gap",),
        ("section", "8. 컬럼 정의 (이슈 관리 시트)"),
        ("note", "이슈 기록의 정본은 test-case/{프로젝트}-issues.json입니다. TC 설계 원본과 "
                 "파일을 나눈 이유는 크기가 아니라 생명주기입니다 — TC 세트는 확정되면 "
                 "고정되지만 이슈는 실행하면서 계속 늘고 상태가 바뀌므로, 한 파일에 두면 "
                 "이슈 갱신 때마다 설계 원본이 변경되어 확정 시점이 흐려집니다."),
        H("컬럼", "내용", "채우는 규칙"),
        R("프로젝트 ID", "프로젝트 식별자", "SUT명 기준. 목록 시트 참조"),
        R("Issue No.", "발생한 이슈의 식별자",
          "발생한 이슈를 등록하며 부여합니다 — {프로젝트}-{생성 번호순} (1 > 2 > 3 > …). "
          "TC 수행 중 발생한 이슈라면 이 번호를 Test Case 시트의 Issue No.에 적어 잇습니다"),
        R("이슈 상태", "이슈의 현재 상태",
          "Open: 열린 상태(재현됨) / In Progress: 담당자가 수정 진행 중 / Resolved: 해결됨 / "
          "Reopen: 해결된 이슈가 재현되어 다시 엶 / Closed: 재현되지 않아 닫음"),
        R("요약(Summary)", "이슈 제목",
          "재현 경로 요약 — 문단 구분은 \">\", 마지막은 무엇이 잘못됐는지 수동태로.\n"
          "ex) rules/ 폴더 이동 > A.md 실행 시, 크래시 발생됨"),
        R("설명", "재현 방법 상세",
          "Pre-condition(사전 조건) / Reproduce Step(재현 스텝) / Actual Result(실제 결과, ~됨) / "
          "Expected Result(기대 결과, ~되어야 함) / QA Comment(참고) 순서로 작성"),
        R("우선순위", "이슈 심각도",
          "High: 크래시·진입 불가 등 치명 / Medium: 진입은 되나 기능 미동작 / Low: 문구·표기 오류"),
        R("빈도", "발생 빈도",
          "Always: 항상 / Often: 종종(70% 이상) / Sometimes: 가끔(70% 미만) / Once: 1회만"),
        R("영향 받는 버전", "이슈가 재현되는 빌드",
          "{테스트환경}_{개발목표버전}_{스프린트}_{확인버전} 네 토막.\n"
          "테스트환경=어디서 확인했는가(PC웹) / 개발목표버전=무엇을 향해 개발 중인가(Ver1.0) / "
          "스프린트=그 목표 버전의 주 목표(Dev, RT1=Release Train 1차) / "
          "확인버전=그 시점의 RC 빌드(RC=Release Candidate).\n"
          "ex) PC웹_Ver1.0_Dev_RC1 · PC웹_Ver1.1_RT1_RC1. 목록 시트 참조(수정 버전과 공용).\n"
          "확인버전은 검증 대상 빌드가 바뀔 때마다 올리며, SUT를 직접 만드는 프로젝트는 그 값을 "
          "SUT 푸터에 표시합니다"),
        R("수정 버전", "이슈가 수정된 버전", "영향 받는 버전과 같은 목록을 사용합니다"),
        R("해결책", "수정 방향",
          "Fixed: 코드 수정으로 대응 / Won't Do: 환경·기술 지원 불가로 미대응 / "
          "Won't Fix: 수정 필요하나 지금은 안 함 / Duplicate: 동일 건 존재 / "
          "Incomplete: 이슈 자체가 잘못됨 / Cannot Reproduce: 재현 불가 / "
          "Done: 기획·정책 결정으로 이슈 아님 확정 / Declined(QA Only): QA선에서 이슈로 보지 않음"),
        R("환경", "발생 디바이스", "PC웹 / Android / iOS"),
        R("레이블", "발생 영역",
          "기능 트리 1-Depth 영역명 — 목록 시트 참조. 트리 개정 시 목록만 갱신하고 기존 행 값은 "
          "소급 변경하지 않습니다"),
        R("보고자", "이슈 등록자", "목록 참조"),
        R("담당자", "이슈 수정 담당자", "목록 참조"),
        R("관측자", "함께 팔로우할 인원",
          "드롭다운 없음 — 쉼표 구분 자유 입력(멀티 선택은 xlsx 표준 기능에 없음)"),
        R("첨부파일", "재현 증적", "재현 영상·이미지의 경로 또는 링크"),
        R("스프린트", "QA 진행 단위",
          "영향 받는 버전에서 확인버전(RC)을 뺀 앞 세 토막 — {테스트환경}_{개발목표버전}_{스프린트}. "
          "스프린트 하나가 RC1·RC2·RC3에 걸치므로 스프린트는 사이클을, 버전은 그 안의 빌드를 "
          "가리킵니다. 발견 단계도 이 칸으로 갈립니다 — ex) PC웹_Ver1.0_Dev(개발 단계 자체 발견) / "
          "PC웹_Ver1.0_RT1(QA 사이클 검출). 목록 참조"),
        R("이슈 등록일", "최초 등록일", "YYYY-MM-DD. 이후 변경하지 않습니다"),
        R("이슈 최종 수정일자", "내용 최종 수정일", "YYYY-MM-DD. 요약·설명 등 내용이 바뀔 때 갱신"),
        R("이슈 상태 최종 변경일자", "상태 최종 변경일", "YYYY-MM-DD. 이슈 상태가 바뀔 때 갱신"),
        ("gap",),
        ("section", "9. 목록 시트 규칙"),
        H("항목", "규칙"),
        R("지위", "드롭다운 참조의 정본 — 전 시트의 데이터 검증이 이 시트를 참조합니다"),
        R("노출", "숨기지 않고 맨 뒤 배치 — 값 체계 자체가 전시물입니다"),
        R("갱신", "항목 추가는 행 추가로. 명칭 변경은 목록만 교체하고 기존 데이터 행은 소급 변경하지 "
                "않습니다(검증은 신규 입력에만 작동)"),
        R("레이블 동기", "기능 트리 개정 시 레이블 목록을 함께 갱신합니다(정본 수정 체크리스트 연동)"),
    ]

    r = 2
    for item in SPEC:
        kind = item[0]
        if kind == "gap":
            r += 1
            continue
        if kind == "title":
            ws.cell(r, 2, item[1]).font = dfont(13, bold=True)
        elif kind == "note":
            x = ws.cell(r, 2, item[1])
            x.font = Font(name=FONT, size=10, color=NOTE_TEXT)
        elif kind == "section":
            ws.cell(r, 2, item[1]).font = dfont(12, bold=True)
        elif kind == "head":
            for i, v in enumerate(item[1:]):
                x = ws.cell(r, 2 + i, v)
                x.font = hfont(11)
                x.fill = HDR
                x.alignment = Alignment(horizontal="center", vertical="center")
                x.border = BOX
        elif kind == "row":
            for i, v in enumerate(item[1:]):
                # '='로 시작하는 문구는 수식으로 저장되어 파일을 깨뜨린다(함정 3).
                # 명세서의 수식 원형은 '=' 없이 적고 본문에서 안내한다.
                assert not (isinstance(v, str) and v.startswith("=")), \
                    f"명세서 셀이 '='로 시작합니다: {v[:40]}"
                x = ws.cell(r, 2 + i, v)
                x.font = dfont(11)
                x.fill = WHITE
                x.alignment = Alignment(vertical="top", wrap_text=True)
                x.border = BOX
        r += 1

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 90
    unfreeze(ws)


if __name__ == "__main__":
    main()
