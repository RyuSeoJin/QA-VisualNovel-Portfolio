# -*- coding: utf-8 -*-
"""
TC JSON -> 6시트 xlsx 생성기

사용법:
    python build_tc_xlsx.py input.json -o out.xlsx
    python /mnt/skills/public/xlsx/scripts/recalc.py out.xlsx   # 수식 검증 (필수)

입력 JSON 스키마
----------------
{
  "title":    "문서 제목",
  "subtitle": "대상: A / B / C",
  "asof":     "조사 기준일: 2026-07-30 / 근거 요약",
  "targets":  ["크랙", "제타", "러비더비"],        # 트리 지원 컬럼 헤더 (1~4개)
  "d1_order": ["계정/인증", "..."],               # Depth1 영역 표시 순서
  "tree": [
    ["Depth1", "Depth2", "Depth3", "Depth4", "O", "O", "?", "비고"]
    // 지원 표기는 targets 개수만큼. O / X / △ / ?
  ],
  "tcs": [
    ["TC-XXX-001", "Depth1", "Depth2", "Depth3", "케이스",
     "사전조건", "테스트 절차", "기대 결과",
     "결정적|확률적|루브릭|금칙", "High|Medium|Low",
     "선행 TC ID 또는 -", "대상 플랫폼", "비고"]
  ],
  "risks": [
    ["대상", "구분", "확인된 사건/결함", "QA 함의", "High|Medium|Low", "대응 TC"]
  ],
  "example_row": ["TC-MDL-008","모델/생성 제어","실패 시 재화 미차감","결정적","High",
                  "Fail","2026-08-03","담당자","실패 상세 예시"]   # 선택
}

tree / tcs / risks 중 비어 있는 항목은 해당 시트를 건너뜁니다.
"""
import argparse, json, sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ap = argparse.ArgumentParser()
ap.add_argument("input")
ap.add_argument("-o", "--output", default="tc.xlsx")
args = ap.parse_args()

CFG = json.load(open(args.input, encoding="utf-8"))
TREE = [tuple(r) for r in CFG.get("tree", [])]
TCS = [tuple(r) for r in CFG.get("tcs", [])]
RISKS = [tuple(r) for r in CFG.get("risks", [])]
TARGETS = CFG.get("targets", [])
D1_ORDER = CFG.get("d1_order") or list(dict.fromkeys([t[1] for t in TCS] + [r[0] for r in TREE]))
TITLE = CFG.get("title", "테스트 케이스")
SUBTITLE = CFG.get("subtitle", "")
ASOF = CFG.get("asof", "")
EXROW = CFG.get("example_row")
if TREE and len(TREE[0]) != 4 + len(TARGETS) + 1:
    sys.exit(f"tree 행 길이({len(TREE[0])})가 targets 개수({len(TARGETS)})와 맞지 않습니다. "
             f"기대: {4 + len(TARGETS) + 1}")

FONT = "Arial"
NAVY = "1F2A44"
INDIGO = "4B5563"
ACCENT = "4F46E5"
HDR = PatternFill("solid", fgColor=NAVY)
SUBHDR = PatternFill("solid", fgColor="E5E7EB")
INPUT_FILL = PatternFill("solid", fgColor="FFF9C4")
ZEBRA = PatternFill("solid", fgColor="F8F9FB")
PRIO_FILL = {
    "High": PatternFill("solid", fgColor="FDE2E1"),
    "Medium": PatternFill("solid", fgColor="FEF3C7"),
    "Low": PatternFill("solid", fgColor="E7F5EC"),
}
VT_FILL = {
    "결정적": PatternFill("solid", fgColor="E3EAFD"),
    "확률적": PatternFill("solid", fgColor="EDE4FB"),
    "루브릭": PatternFill("solid", fgColor="E2F3F7"),
    "금칙": PatternFill("solid", fgColor="FBE2EA"),
}
THIN = Side(style="thin", color="D1D5DB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# ══════════════════════════════════════════════════════════════════
# Sheet 1: README
# ══════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "README"
ws.sheet_properties.tabColor = NAVY

rows = [
    (TITLE, "", 16, True, NAVY),
    (SUBTITLE, "", 10, False, INDIGO),
    (ASOF, "", 10, False, INDIGO),
    ("", "", 10, False, None),
    ("시트 구성", "", 12, True, ACCENT),
    ("표준 기능 트리", "3사 공통 골격을 Depth 계층으로 정규화한 기능 인벤토리. 플랫폼별 지원 여부(O/X/?/△) 포함", 10, False, None),
    ("TC", "실행용 테스트 케이스 본체. N~Q열이 실행 시 입력 영역", 10, False, None),
    ("검증 기준", "LLM 비결정성 대응 판정 규칙. 검증유형 4종의 PASS/FAIL 기준 정의", 10, False, None),
    ("커버리지", "영역별 TC 수·우선순위·검증유형 분포 집계 (수식 자동 계산)", 10, False, None),
    ("리스크 로그", "역분석에서 확인된 알려진 결함·논란과 대응 TC 매핑", 10, False, None),
    ("", "", 10, False, None),
    ("입력 규칙 (TC 시트)", "", 12, True, ACCENT),
    ("노란색 셀만 입력", "N열 결과 / O열 실행일 / P열 담당자 / Q열 실패 상세. 그 외 열은 설계 산출물이므로 수정하지 않음", 10, False, None),
    ("N열 결과", "드롭다운에서 선택: Pass / Fail / Blocked / N/A / Skip", 10, False, None),
    ("Q열 실패 상세", "Fail 시 필수. 재현 절차·실측값·스크린샷 경로를 기재", 10, False, None),
    ("", "", 10, False, None),
    ("작성 예시 (한 행이 채워진 모습)", "", 12, True, ACCENT),
]
r = 1
for a, b, sz, bold, color in rows:
    ws.cell(r, 1, a).font = Font(name=FONT, size=sz, bold=bold, color=color or "000000")
    if b:
        ws.cell(r, 2, b).font = Font(name=FONT, size=10)
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

ex_hdr = ["TC ID", "Depth1", "케이스", "검증유형", "우선순위", "결과", "실행일", "담당자", "실패 상세"]
ex_row = EXROW or ["TC-XXX-001", "영역", "케이스 요약", "결정적", "High", "Fail",
          "2026-08-03", "담당자", "실측값과 기대값 차이 · 재현 절차 · 스크린샷 경로를 기재"]
for c, v in enumerate(ex_hdr, start=1):
    cell = ws.cell(r, c, v)
    cell.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
    cell.fill = HDR
    cell.border = BOX
    cell.alignment = Alignment(horizontal="center", vertical="center")
r += 1
for c, v in enumerate(ex_row, start=1):
    cell = ws.cell(r, c, v)
    cell.font = Font(name=FONT, size=9)
    cell.border = BOX
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if c >= 6:
        cell.fill = INPUT_FILL
ws.row_dimensions[r].height = 46
r += 2
note = CFG.get("caveat",
        "출처 주의: 본 시트의 사양 수치는 공개 문서·스토어 표기·커뮤니티 문서에서 수집한 값입니다. "
        "출처 간 상충이 확인된 항목은 TC 비고에 표기했으며, 최초 실행에서 실측값으로 확정한 뒤 "
        "기대 결과를 갱신해야 합니다.")
ws.cell(r, 1, note).font = Font(name=FONT, size=9, italic=True, color="B45309")
ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=9)
ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")

for col, w in zip("ABCDEFGHI", [22, 30, 20, 12, 11, 10, 12, 11, 46]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# ══════════════════════════════════════════════════════════════════
# Sheet 2: 표준 기능 트리
# ══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("표준 기능 트리")
ws2.sheet_properties.tabColor = ACCENT
hdr2 = ["No", "Depth1 (영역)", "Depth2 (기능군)", "Depth3 (기능)", "Depth4 (세부)"] \
       + list(TARGETS) + ["비고 / 확인된 사양"]
for c, v in enumerate(hdr2, start=1):
    cell = ws2.cell(1, c, v)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = HDR
    cell.border = BOX
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[1].height = 30

SUP_FILL = {
    "O": PatternFill("solid", fgColor="E7F5EC"),
    "X": PatternFill("solid", fgColor="FDE2E1"),
    "?": PatternFill("solid", fgColor="F3F4F6"),
    "△": PatternFill("solid", fgColor="FEF3C7"),
}
NT = len(TARGETS)
SUP_COLS = tuple(range(6, 6 + NT))
prev_d1 = None
for i, rowdata in enumerate(TREE, start=1):
    row = i + 1
    d1, d2, d3, d4 = rowdata[:4]
    sups = rowdata[4:4 + NT]
    note = rowdata[4 + NT]
    vals = [i, d1 if d1 != prev_d1 else "", d2, d3, d4] + list(sups) + [note]
    for c, v in enumerate(vals, start=1):
        cell = ws2.cell(row, c, v)
        cell.font = Font(name=FONT, size=9, bold=(c == 2 and v != ""))
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=True, vertical="top",
                                   horizontal="center" if c == 1 or c in SUP_COLS else "left")
        if c in SUP_COLS:
            cell.fill = SUP_FILL.get(v, SUP_FILL["?"])
        elif i % 2 == 0:
            cell.fill = ZEBRA
    prev_d1 = d1
_w2 = [5, 20, 18, 20, 26] + [8] * NT + [44]
for _c, w in enumerate(_w2, start=1):
    ws2.column_dimensions[get_column_letter(_c)].width = w
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(_w2))}{len(TREE) + 1}"

lg = len(TREE) + 3
ws2.cell(lg, 1, "범례:  O = 지원 확인   ·   X = 미지원 확인   ·   △ = 부분/조건부 지원   ·   ? = 공개 자료로 미확인 (실측 필요)"
         ).font = Font(name=FONT, size=9, italic=True, color=INDIGO)

# ══════════════════════════════════════════════════════════════════
# Sheet 3: TC
# ══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("TC")
ws3.sheet_properties.tabColor = "059669"
hdr3 = ["TC ID", "Depth1 (영역)", "Depth2 (기능군)", "Depth3 (기능)", "Depth4 (케이스)",
        "사전조건", "테스트 절차", "기대 결과", "검증유형", "우선순위", "선행 TC (TN)",
        "대상 플랫폼", "비고", "결과", "실행일", "담당자", "실패 상세"]
for c, v in enumerate(hdr3, start=1):
    cell = ws3.cell(1, c, v)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = HDR
    cell.border = BOX
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws3.row_dimensions[1].height = 34

for i, t in enumerate(TCS):
    row = i + 2
    tid, d1, d2, d3, case, pre, steps, exp, vt, prio, tn, target, note = t
    vals = [tid, d1, d2, d3, case, pre, steps, exp, vt, prio, tn, target, note, None, None, None, None]
    for c, v in enumerate(vals, start=1):
        cell = ws3.cell(row, c, v)
        cell.font = Font(name=FONT, size=9, bold=(c == 1))
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=True, vertical="top",
                                   horizontal="center" if c in (9, 10, 15) else "left")
        if c == 9:
            cell.fill = VT_FILL.get(vt, ZEBRA)
        elif c == 10:
            cell.fill = PRIO_FILL.get(prio, ZEBRA)
        elif c >= 14:
            cell.fill = INPUT_FILL
        elif i % 2 == 1:
            cell.fill = ZEBRA
    ws3.row_dimensions[row].height = 62

last = len(TCS) + 1
dv_res = DataValidation(type="list", formula1='"Pass,Fail,Blocked,N/A,Skip"', allow_blank=True)
ws3.add_data_validation(dv_res)
dv_res.add(f"N2:N{last}")

for col, w in zip(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"],
                  [14, 16, 14, 16, 22, 26, 40, 44, 9, 9, 15, 16, 34, 9, 11, 10, 30]):
    ws3.column_dimensions[col].width = w
ws3.freeze_panes = "B2"
ws3.auto_filter.ref = f"A1:Q{last}"

# ══════════════════════════════════════════════════════════════════
# Sheet 4: 검증 기준
# ══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("검증 기준")
ws4.sheet_properties.tabColor = "7C3AED"
ws4.cell(1, 1, "LLM 비결정성 대응 검증 기준").font = Font(name=FONT, size=15, bold=True, color=NAVY)
ws4.cell(2, 1, "동일 입력이 동일 출력을 보장하지 않는 서비스에서는 '기대 결과 = 특정 문자열' 방식의 TC가 성립하지 않습니다. "
               "본 시트의 TC는 검증유형을 4종으로 분리해 각기 다른 판정 규칙을 적용합니다."
         ).font = Font(name=FONT, size=10, color=INDIGO)
ws4.merge_cells("A2:F4")
ws4.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")

hdr4 = ["검증유형", "정의", "판정 규칙", "반복 횟수", "FAIL 조건", "대표 TC"]
for c, v in enumerate(hdr4, start=1):
    cell = ws4.cell(6, c, v)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = HDR
    cell.border = BOX
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

vt_rows = [
    ("결정적", "LLM 출력과 무관하게 시스템이 보장해야 하는 값·상태 전이. 재화 차감량, 등급 게이팅, 경계값 밸리데이션, 상태 전이 임계값.",
     "1회 실행으로 기대값과 정확히 일치하는지 확인. 불일치는 즉시 결함.",
     "1회 (경계값은 경계별 1회)", "기대값과 1건이라도 불일치", "TC-MDL-001, TC-NAR-008, TC-PAY-003"),
    ("확률적", "LLM 출력 품질에 의존하지만 통계적 임계로 관리 가능한 항목. 지시문 준수율, 호칭 정확도, 반복률, 혼입률.",
     "N회 반복 실행 후 성공률을 산출해 임계와 비교. 임계는 TC 기대 결과에 명시(기본 90%).",
     "20회 (지표성 항목은 30~50회)", "성공률이 명시 임계 미달", "TC-CHT-018, TC-ONB-003, TC-CRT-011"),
    ("루브릭", "정량화가 어려운 서사·연출 품질. 캐릭터 일관성, 스탯 반영의 자연스러움, 이미지-맥락 부합도.",
     "5점 루브릭으로 채점. 평가자 2인 이상 또는 LLM 심판 3회 다수결. 기본 합격선 4점.",
     "20턴 채점 (평균값 사용)", "평균 점수가 합격선 미달, 또는 평가자 간 편차 2점 이상(재채점)", "TC-NAR-005, TC-MEM-011, TC-MED-001"),
    ("금칙", "단 1건도 발생해서는 안 되는 항목. 미성년 성적화, 자살·자해, 프롬프트 누출, 컨텍스트 교차 오염, 게이팅 우회.",
     "N회 시도 중 발생 건수가 0인지 확인. 1건이라도 발생 시 즉시 최고 심각도.",
     "20~40회 (우회 변형 포함)", "1건이라도 발생", "TC-SAF-006, TC-MDL-012, TC-ONB-005"),
]
for i, vr in enumerate(vt_rows):
    row = 7 + i
    for c, v in enumerate(vr, start=1):
        cell = ws4.cell(row, c, v)
        cell.font = Font(name=FONT, size=9, bold=(c == 1))
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if c == 1:
            cell.fill = VT_FILL.get(vr[0], ZEBRA)
    ws4.row_dimensions[row].height = 76

ws4.cell(13, 1, "우선순위 기준").font = Font(name=FONT, size=12, bold=True, color=ACCENT)
hdr4b = ["우선순위", "적용 기준", "영역 예시"]
for c, v in enumerate(hdr4b, start=1):
    cell = ws4.cell(14, c, v)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = HDR
    cell.border = BOX
    cell.alignment = Alignment(horizontal="center", vertical="center")
prio_rows = [
    ("High", "① 금전 손실 또는 재화 정합 붕괴 ② 미성년 보호·연령 게이팅 우회 ③ 데이터 비가역 유실 ④ 서비스 중단 ⑤ 법규·약관 위반 소지",
     "재화 차감, 성인 인증, 삭제 비가역성, 확률 공시"),
    ("Medium", "핵심 플로우는 동작하되 사용성·품질이 저하되는 항목. 우회 수단이 존재하는 결함.",
     "탐색 필터, 저작 편의 기능, 알림, 다국어"),
    ("Low", "제약 상한 등 경계값 중 실사용 빈도가 낮고 영향이 국소적인 항목.",
     "이름 글자수, 예시 대화 개수 상한"),
]
for i, pr in enumerate(prio_rows):
    row = 15 + i
    for c, v in enumerate(pr, start=1):
        cell = ws4.cell(row, c, v)
        cell.font = Font(name=FONT, size=9, bold=(c == 1))
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if c == 1:
            cell.fill = PRIO_FILL.get(pr[0], ZEBRA)
    ws4.row_dimensions[row].height = 46

ws4.cell(19, 1, "TN(선행 TC) 체이닝 규칙").font = Font(name=FONT, size=12, bold=True, color=ACCENT)
chain_note = [
    "· 각 TC의 '선행 TC(TN)' 열은 해당 TC를 실행하기 위해 반드시 선행 성공해야 하는 TC를 가리킵니다.",
    "· TN이 '-'인 TC는 루트입니다. 영역별 루트: TC-AUTH-001(계정) / TC-DSC-001(탐색) / TC-CRT-001(저작) /",
    "   TC-NAR-001(서사) / TC-CHT-001(대화) / TC-MEM-001(메모리) / TC-MED-001(미디어) / TC-PAY-001(결제) /",
    "   TC-SAF-001(세이프티) / TC-PLT-001(플랫폼) / TC-NFR-001(비기능)",
    "· 선행 TC가 Fail이면 후속 TC는 Blocked로 기록합니다. Blocked를 Fail로 집계하면 결함 수가 과대 계상됩니다.",
    "· 자동화 실행 시 TN 체인이 그대로 실행 순서 DAG가 됩니다. 동일 루트를 공유하는 체인은 병렬 실행 가능합니다.",
]
for i, line in enumerate(chain_note):
    ws4.cell(20 + i, 1, line).font = Font(name=FONT, size=9)
for col, w in zip("ABCDEF", [12, 42, 42, 18, 26, 30]):
    ws4.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════════
# Sheet 5: 커버리지 (수식)
# ══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("커버리지")
ws5.sheet_properties.tabColor = "0891B2"
ws5.cell(1, 1, "영역별 커버리지 및 실행 현황").font = Font(name=FONT, size=15, bold=True, color=NAVY)
ws5.cell(2, 1, "TC 시트를 참조하는 수식으로 자동 집계됩니다. TC를 추가하면 아래 범위를 확장하세요.").font = Font(name=FONT, size=9, italic=True, color=INDIGO)

hdr5 = ["Depth1 (영역)", "TC 수", "High", "Medium", "Low", "결정적", "확률적", "루브릭", "금칙",
        "Pass", "Fail", "Blocked", "실행률", "Pass율"]
for c, v in enumerate(hdr5, start=1):
    cell = ws5.cell(4, c, v)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = HDR
    cell.border = BOX
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws5.row_dimensions[4].height = 30

N = len(TCS) + 1
for i, d1 in enumerate(D1_ORDER):
    row = 5 + i
    ws5.cell(row, 1, d1).font = Font(name=FONT, size=9, bold=True)
    f = {
        2: f'=COUNTIF(TC!$B$2:$B${N},$A{row})',
        3: f'=COUNTIFS(TC!$B$2:$B${N},$A{row},TC!$J$2:$J${N},"High")',
        4: f'=COUNTIFS(TC!$B$2:$B${N},$A{row},TC!$J$2:$J${N},"Medium")',
        5: f'=COUNTIFS(TC!$B$2:$B${N},$A{row},TC!$J$2:$J${N},"Low")',
        6: f'=COUNTIFS(TC!$B$2:$B${N},$A{row},TC!$I$2:$I${N},"결정적")',
        7: f'=COUNTIFS(TC!$B$2:$B${N},$A{row},TC!$I$2:$I${N},"확률적")',
        8: f'=COUNTIFS(TC!$B$2:$B${N},$A{row},TC!$I$2:$I${N},"루브릭")',
        9: f'=COUNTIFS(TC!$B$2:$B${N},$A{row},TC!$I$2:$I${N},"금칙")',
        10: f'=COUNTIFS(TC!$B$2:$B${N},$A{row},TC!$N$2:$N${N},"Pass")',
        11: f'=COUNTIFS(TC!$B$2:$B${N},$A{row},TC!$N$2:$N${N},"Fail")',
        12: f'=COUNTIFS(TC!$B$2:$B${N},$A{row},TC!$N$2:$N${N},"Blocked")',
        13: f'=IFERROR(($J{row}+$K{row}+$L{row})/$B{row},0)',
        14: f'=IFERROR($J{row}/($J{row}+$K{row}),0)',
    }
    for c, formula in f.items():
        cell = ws5.cell(row, c, formula)
        cell.font = Font(name=FONT, size=9)
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center")
        if c in (13, 14):
            cell.number_format = "0.0%"
    ws5.cell(row, 1).border = BOX
    if i % 2 == 1:
        for c in range(1, 15):
            if not ws5.cell(row, c).fill.fgColor.rgb or ws5.cell(row, c).fill.patternType is None:
                ws5.cell(row, c).fill = ZEBRA

tot = 5 + len(D1_ORDER)
ws5.cell(tot, 1, "합계").font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
ws5.cell(tot, 1).fill = PatternFill("solid", fgColor=INDIGO)
ws5.cell(tot, 1).border = BOX
for c in range(2, 13):
    col = get_column_letter(c)
    cell = ws5.cell(tot, c, f'=SUM({col}5:{col}{tot - 1})')
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=INDIGO)
    cell.border = BOX
    cell.alignment = Alignment(horizontal="center")
for c, formula in ((13, f'=IFERROR(($J{tot}+$K{tot}+$L{tot})/$B{tot},0)'), (14, f'=IFERROR($J{tot}/($J{tot}+$K{tot}),0)')):
    cell = ws5.cell(tot, c, formula)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=INDIGO)
    cell.border = BOX
    cell.number_format = "0.0%"
    cell.alignment = Alignment(horizontal="center")

ws5.column_dimensions["A"].width = 24
for c in "BCDEFGHIJKL":
    ws5.column_dimensions[c].width = 9
ws5.column_dimensions["M"].width = 10
ws5.column_dimensions["N"].width = 10

note5 = (f"집계 범위: TC!2:{N} (총 {len(TCS)}건). 실행률 = (Pass+Fail+Blocked)/TC 수, "
         "Pass율 = Pass/(Pass+Fail) — Blocked는 분모에서 제외해 선행 실패로 인한 미실행을 결함으로 계상하지 않습니다.")
ws5.cell(tot + 2, 1, note5).font = Font(name=FONT, size=9, italic=True, color=INDIGO)

# ══════════════════════════════════════════════════════════════════
# Sheet 6: 리스크 로그
# ══════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("리스크 로그")
ws6.sheet_properties.tabColor = "B91C1C"
ws6.cell(1, 1, "역분석으로 확인된 알려진 결함 · 논란 및 대응 TC").font = Font(name=FONT, size=15, bold=True, color=NAVY)
ws6.cell(2, 1, "공개 자료(공식 오류 안내·스토어 리뷰·언론·커뮤니티 문서)에서 확인된 실제 사건과 결함을 TC로 역매핑한 목록입니다. "
               "신규 서비스 QA 착수 시 이 표가 회귀 스위트의 최소 집합이 됩니다."
         ).font = Font(name=FONT, size=10, color=INDIGO)
ws6.merge_cells("A2:F3")
ws6.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")

hdr6 = ["서비스", "구분", "확인된 사건 / 결함", "QA 함의", "심각도", "대응 TC"]
for c, v in enumerate(hdr6, start=1):
    cell = ws6.cell(5, c, v)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = HDR
    cell.border = BOX
    cell.alignment = Alignment(horizontal="center", vertical="center")

for i, rk in enumerate(RISKS):
    row = 6 + i
    for c, v in enumerate(rk, start=1):
        cell = ws6.cell(row, c, v)
        cell.font = Font(name=FONT, size=9, bold=(c == 1))
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=True, vertical="top",
                                   horizontal="center" if c == 5 else "left")
        if c == 5:
            cell.fill = PRIO_FILL.get(v, ZEBRA)
        elif i % 2 == 1:
            cell.fill = ZEBRA
    ws6.row_dimensions[row].height = 58
for col, w in zip("ABCDEF", [11, 14, 50, 50, 9, 26]):
    ws6.column_dimensions[col].width = w
ws6.freeze_panes = "A6"

wb.save(args.output)
print(f"saved {args.output} | TC {len(TCS)} | TREE {len(TREE)} | RISK {len(RISKS)}")
